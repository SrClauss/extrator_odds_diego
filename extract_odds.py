#!/usr/bin/env python3
"""
Script de extração de odds para Excel usando Selenium
Automatiza login no UltraVirtual, navega até Bet365 e extrai odds para Excel.
"""

import re
import time
import subprocess
import sys
import unicodedata
from datetime import datetime
from collections import defaultdict

#               Função para instalar pacotes caso não existam
def instalar_se_necessario(pacote, import_name=None):
    """Instala pacote se não estiver disponível"""
    if import_name is None:
        import_name = pacote
    try:
        __import__(import_name)
    except ImportError:
        print(f"📦 Instalando {pacote}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pacote, "-q"])

# Verificar e instalar dependências
instalar_se_necessario("beautifulsoup4", "bs4")
instalar_se_necessario("openpyxl")
instalar_se_necessario("selenium")
instalar_se_necessario("webdriver-manager", "webdriver_manager")

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


def extrair_odds_do_html(html_source):
    """
    Extrai odds do HTML fornecido.
    Retorna lista de dicts com 'odd' e 'cor'.
    Busca apenas divs com classe br-tb e background verde (#28a745) ou vermelho (#dc3545).
    O valor da odd está no div interno com color: rgb(255, 215, 0).
    """
    soup = BeautifulSoup(html_source, 'html.parser')
    odds_data = []

    # Buscar apenas as células do grid (classe br-tb)
    celulas = soup.find_all('div', class_='br-tb')

    for celula in celulas:
        style = celula.get('style', '')

        # Apenas células verdes ou vermelhas
        if '#dc3545' not in style and '#28a745' not in style:
            continue

        cor = 'Vermelho' if '#dc3545' in style else 'Verde'

        # O valor da odd fica no div interno com cor dourada
        div_valor = celula.find('div', style=lambda s: s and 'color: rgb(255, 215, 0)' in s)
        if not div_valor:
            continue

        text_content = div_valor.get_text(strip=True)
        if not text_content:
            continue

        match = re.search(r'\d+\.?\d*', text_content)
        if match:
            odds_data.append({
                'odd': match.group(),
                'cor': cor
            })

    return odds_data


def agregar_odds(odds_data):
    """
    Agrega odds idênticas e conta ocorrências por cor.
    Retorna lista de dicts ordenada por valor numérico.
    """
    resumo = defaultdict(lambda: {'verde': 0, 'vermelho': 0})
    
    for item in odds_data:
        odd_val = float(item['odd'])
        cor_type = 'verde' if item['cor'] == 'Verde' else 'vermelho'
        resumo[odd_val][cor_type] += 1
    
    # Converter para lista e ordenar por valor numérico
    dados_planilha = []
    for odd_val in sorted(resumo.keys()):
        total = resumo[odd_val]['verde'] + resumo[odd_val]['vermelho']
        dados_planilha.append({
            'Odd': str(odd_val),
            'Verde': resumo[odd_val]['verde'],
            'Vermelho': resumo[odd_val]['vermelho'],
            'Total': total
        })
    
    return dados_planilha


def sanitizar_nome(nome):
    """
    Converte nome de competição para slug de arquivo.
    Ex: 'Copa do Mundo' → 'copa_do_mundo'
    """
    nome = unicodedata.normalize('NFKD', nome)
    nome = ''.join(c for c in nome if not unicodedata.combining(c))
    nome = nome.lower()
    nome = re.sub(r'[^a-z0-9]+', '_', nome)
    nome = nome.strip('_')
    return nome


def adicionar_sheet_excel(workbook, dados_planilha, nome_sheet):
    """
    Adiciona uma nova aba (sheet) ao workbook com dados formatados.
    """
    # Criar nova aba com o nome do mercado
    ws = workbook.create_sheet(title=nome_sheet)
    
    # Definir estilos
    header_fill = PatternFill(start_color="0f273d", end_color="0f273d", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    verde_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    vermelho_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalho
    headers = ['Odd', 'Verde', 'Vermelho', 'Total', 'Odd Real', 'Teórica', 'Real - Teórica']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Dados
    for row_idx, item in enumerate(dados_planilha, 2):
        # Coluna Odd (numérico, formato sem zeros desnecessários)
        cell = ws.cell(row=row_idx, column=1)
        cell.value = float(item['Odd'])
        cell.number_format = '0.##'
        cell.alignment = center_alignment

        # Coluna Verde
        cell = ws.cell(row=row_idx, column=2)
        cell.value = item['Verde']
        cell.fill = verde_fill
        cell.alignment = center_alignment

        # Coluna Vermelho
        cell = ws.cell(row=row_idx, column=3)
        cell.value = item['Vermelho']
        cell.fill = vermelho_fill
        cell.alignment = center_alignment

        # Coluna Total
        cell = ws.cell(row=row_idx, column=4)
        cell.value = item['Total']
        cell.alignment = center_alignment

        # Coluna Odd Real = Verde / Total
        cell = ws.cell(row=row_idx, column=5)
        cell.value = f'=IFERROR(B{row_idx}/D{row_idx},0)'
        cell.number_format = '0.00%'
        cell.alignment = center_alignment

        # Coluna Teórica = 1 / Odd
        cell = ws.cell(row=row_idx, column=6)
        cell.value = f'=1/A{row_idx}'
        cell.number_format = '0.00%'
        cell.alignment = center_alignment

        # Coluna Real - Teórica
        cell = ws.cell(row=row_idx, column=7)
        cell.value = f'=E{row_idx}-F{row_idx}'
        cell.number_format = '0.00%'
        cell.alignment = center_alignment

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16


def criar_excel(dados_planilha, nome_arquivo):
    """
    Cria workbook Excel com dados formatados e salva.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Odds"
    
    # Definir estilos
    header_fill = PatternFill(start_color="0f273d", end_color="0f273d", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    verde_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    vermelho_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalho
    headers = ['Odd', 'Verde', 'Vermelho', 'Total']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Dados
    for row_idx, item in enumerate(dados_planilha, 2):
        # Coluna Odd
        cell = ws.cell(row=row_idx, column=1)
        cell.value = item['Odd']
        cell.alignment = center_alignment
        
        # Coluna Verde
        cell = ws.cell(row=row_idx, column=2)
        cell.value = item['Verde']
        cell.fill = verde_fill
        cell.alignment = center_alignment
        
        # Coluna Vermelho
        cell = ws.cell(row=row_idx, column=3)
        cell.value = item['Vermelho']
        cell.fill = vermelho_fill
        cell.alignment = center_alignment
        
        # Coluna Total
        cell = ws.cell(row=row_idx, column=4)
        cell.value = item['Total']
        cell.alignment = center_alignment
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    
    # Salvar arquivo
    wb.save(nome_arquivo)
    return nome_arquivo


def capturar_html_selenium(driver):
    """
    Captura o HTML do grid usando o elemento já carregado.
    """
    try:
        # Pegar o elemento do grid com a classe br-tb
        grid_element = driver.find_element(By.CLASS_NAME, 'br-tb')
        
        # Capturar o HTML completo da página
        html_grid = driver.page_source
        
        return html_grid
    except Exception as e:
        print(f"❌ Erro ao capturar grid: {e}")
        return None


def obter_escolha_mercado():
    """
    Exibe menu de mercados e retorna o mercado selecionado pelo usuário.
    Não requer Selenium - apenas input do usuário.
    """
    mercados = [
        "Resultado Final",
        "Over Gols",
        "Under Gols",
        "Para o Time Marcar Sim/Não",
        "Resultado Correto - FT",
        "Primeiro Marcador de Gol",
        "Resultado Correto - Grupo",
        "Resultado/Para Ambos Times Marcarem",
        "Dupla Hipótese",
        "Intervalo/Final do Jogo",
        "Total de Gols Exatos",
        "Intervalo - Resultado",
        "Resultado Correto - HT",
        "Margem de Vitória",
        "Time a Marcar",
        "Time - Gols",
        "Handicap - Resultado"
    ]
    
    print("\n" + "=" * 60)
    print(" SELECIONE UM MERCADO")
    print("=" * 60 + "\n")
    
    for i, mercado in enumerate(mercados, 1):
        print(f" {i:2d}. {mercado}")
    
    print("\n")
    while True:
        try:
            opcao = int(input("Digite o número do mercado desejado: ").strip())
            if 1 <= opcao <= len(mercados):
                mercado_selecionado = mercados[opcao - 1]
                print(f"\n✅ Mercado selecionado: {mercado_selecionado}\n")
                return mercado_selecionado
            else:
                print("❌ Opção inválida! Digite um número da lista.")
        except ValueError:
            print("❌ Digite um número válido!")


def escolher_mercado(driver, mercado_selecionado):
    """
    Aplica a seleção de mercado via Selenium usando a escolha do usuário.
    """
    from selenium.webdriver.support.ui import Select
    
    print(f"🔄 Aplicando filtro: {mercado_selecionado}...")
    try:
        select_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.TAG_NAME, "select"))
        )
        select = Select(select_element)
        select.select_by_visible_text(mercado_selecionado)
        time.sleep(1)
        print(f"✓ Filtro aplicado com sucesso!")
        return mercado_selecionado
    except Exception as e:
        print(f"⚠️  Erro ao selecionar mercado: {e}")
        print("Continuando com mercado padrão...")
        return mercado_selecionado


def listar_competicoes():
    """
    Retorna a lista fixa de competições disponíveis no Bet365 do UltraVirtual.
    Baseada nos ícones do menu: world, euro, premier, super, express.
    """
    base = "https://ultravirtual.com.br"
    return [
        {'nome': 'Copa do Mundo',           'url': f'{base}/dashboard/bet365/world/hourly'},
        {'nome': 'Euro Cup',                'url': f'{base}/dashboard/bet365/euro/hourly'},
        {'nome': 'Premier League',          'url': f'{base}/dashboard/bet365/premier/hourly'},
        {'nome': 'Superliga Sul-Americana', 'url': f'{base}/dashboard/bet365/super/hourly'},
        {'nome': 'Express',                 'url': f'{base}/dashboard/bet365/express/hourly'},
    ]


def fazer_login(driver):
    """
    Faz login no site UltraVirtual.
    """
    login = "diegosantosdeassis5@gmail.com"
    senha = "bomba2022"
    
    print("\n🔐 Fazendo login...")
    
    try:
        # Clicar no botão ENTRAR
        button_ENTRAR = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'ENTRAR')]"))
        )
        button_ENTRAR.click()
        
        # Preencher e-mail
        input1 = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='email' or @type='text']"))
        )
        input1.clear()
        input1.send_keys(login)
        
        # Preencher senha
        input2 = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='password']"))
        )
        input2.clear()
        input2.send_keys(senha)
        input2.submit()
        
        print("✓ Login realizado com sucesso!")
        time.sleep(2)
        return True
        
    except Exception as e:
        print(f"❌ Erro no login: {e}")
        return False


def navegar_ate_odds(driver, url=None):
    """
    Navega até a página de odds de uma competição (Placar FT → Odds) e aguarda o grid carregar.
    Se url não for informado, usa Copa do Mundo como padrão.
    """
    if url is None:
        url = "https://ultravirtual.com.br/dashboard/bet365/world/hourly"
    try:
        print(f"\n🎯 Navegando para: {url}")

        # Ir direto para a URL da competição
        driver.get(url)
        time.sleep(4)  # tempo extra para a página renderizar

        # Clicar em Placar FT
        print("📊 Abrindo Placar FT...")
        try:
            placarFT_button_span = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Placar FT')]"))
            )
            placarFT_button_span.click()
            time.sleep(2)
        except Exception as e:
            print(f"⚠️  Botão 'Placar FT' não encontrado: {e}")
            print("   Continuando sem clicar em Placar FT...")

        # Clicar em Odds
        print("💰 Abrindo grid de Odds...")
        try:
            odds_button_span = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Odds')]"))
            )
            odds_button_span.click()
        except Exception as e:
            print(f"⚠️  Botão 'Odds' não encontrado: {e}")
            print("   Continuando sem clicar em Odds...")

        print("⏳ Aguardando odds carregarem completamente...")
        time.sleep(5)

        # Aguardar até o grid estar presente (não-fatal)
        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CLASS_NAME, "br-tb"))
            )
        except Exception:
            print("⚠️  Grid 'br-tb' não detectado no tempo limite — tentando extrair assim mesmo...")

        print("✓ Navegação concluída!")
        return True

    except Exception as e:
        print(f"❌ Erro fatal na navegação: {e}")
        return False


def clicar_botao_e_extrair(driver, workbook, texto_botao, nome_aba):
    """
    Clica em um botão pelo texto, aguarda o grid atualizar,
    extrai as odds e adiciona uma nova aba ao workbook.
    """
    print(f"\n🔘 Clicando em '{texto_botao}'...")
    try:
        botao = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, f"//button[contains(text(), '{texto_botao}')]"))
        )
        botao.click()
        time.sleep(3)
        print(f"✓ Botão '{texto_botao}' clicado!")
    except Exception as e:
        print(f"❌ Erro ao clicar no botão '{texto_botao}': {e}")
        return

    # Navegar para visualização 24 Horas/Linhas → Células
    try:
        print("🔀 Selecionando visualização 24 Horas/Linhas...")
        linhas = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Linhas')]"))
        )
        linhas.click()
        time.sleep(1)

        select2_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "select"))
        )
        from selenium.webdriver.support.ui import Select as _Select
        select2_ = _Select(select2_element)
        select2_.select_by_visible_text("24 Horas/Linhas")
        time.sleep(1)

        celulas = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Células')]"))
        )
        celulas.click()
        time.sleep(2)
        print("✓ Visualização 24 Horas/Linhas → Células aplicada!")
    except Exception as e:
        print(f"⚠️  Erro ao aplicar visualização Linhas/Células: {e}")

    # Capturar HTML
    print("📥 Capturando HTML do grid...")
    html_grid = capturar_html_selenium(driver)
    if not html_grid:
        print("❌ Falha ao capturar HTML!")
        return

    # Extrair e agregar
    odds_data = extrair_odds_do_html(html_grid)
    print(f"✓ {len(odds_data)} odds extraídas")

    if not odds_data:
        print(f"⚠️  Nenhuma odd encontrada para '{nome_aba}', aba não será criada.")
        return

    dados_planilha = agregar_odds(odds_data)

    total_ocorrencias = len(odds_data)
    total_verde = sum(item['Verde'] for item in dados_planilha)
    total_vermelho = sum(item['Vermelho'] for item in dados_planilha)
    print(f"📈 {len(dados_planilha)} odds únicas | Verde: {total_verde} | Vermelho: {total_vermelho} | Total: {total_ocorrencias}")

    # Adicionar aba ao workbook
    adicionar_sheet_excel(workbook, dados_planilha, nome_aba)
    print(f"✅ Aba adicionada: '{nome_aba}'")


def extrair_competicao(driver, workbook, competicao_url):
    """
    Para uma competição já carregada, executa as 4 extrações de mercado
    e popula o workbook com as abas correspondentes.
    """
    # 1) Over Gols → Over 2.5
    print("\n" + "=" * 60)
    print(" [1/4] Over Gols → Over 2.5")
    print("=" * 60)
    escolher_mercado(driver, "Over Gols")
    clicar_botao_e_extrair(driver, workbook, "Over 2.5", "Over 2.5")

    # 2) Over Gols → Over 3.5
    print("\n" + "=" * 60)
    print(" [2/4] Over Gols → Over 3.5")
    print("=" * 60)
    clicar_botao_e_extrair(driver, workbook, "Over 3.5", "Over 3.5")

    # 3) Total de Gols Exatos → 5+ gols
    print("\n" + "=" * 60)
    print(" [3/4] Total de Gols Exatos → 5+ gols")
    print("=" * 60)
    escolher_mercado(driver, "Total de Gols Exatos")
    clicar_botao_e_extrair(driver, workbook, "5+ gols", "5+ gols")

    # 4) Para o Time Marcar Sim/Não → Ambas Sim
    print("\n" + "=" * 60)
    print(" [4/4] Para o Time Marcar Sim/Não → Ambas Sim")
    print("=" * 60)
    escolher_mercado(driver, "Para o Time Marcar Sim/Não")
    clicar_botao_e_extrair(driver, workbook, "Ambas Sim", "Ambas Sim")


def main():
    """
    Função principal - raspa todas as competições disponíveis no Bet365,
    criando um arquivo Excel separado para cada uma.
    Formato do arquivo: {slug_competicao}_{timestamp}.xlsx
    """
    print("=" * 60)
    print(" EXTRATOR AUTOMÁTICO DE ODDS PARA EXCEL")
    print(" UltraVirtual → Bet365 → Todas as Competições")
    print("=" * 60)

    timestamp = datetime.now().strftime("%d%m%Y_%H%M")

    # Inicializar Selenium
    print("\n🌐 Iniciando navegador Chrome...")
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service)
        print("✓ Chrome iniciado!")
    except Exception as e:
        print(f"❌ Erro ao iniciar Chrome: {e}")
        print("💡 Certifique-se de ter o Google Chrome instalado")
        return

    arquivos_salvos = []

    try:
        # Abrir site e fazer login
        print("\n🔗 Acessando UltraVirtual...")
        driver.get("https://ultravirtual.com.br/")
        time.sleep(2)

        if not fazer_login(driver):
            return

        competicoes = listar_competicoes()
        print(f"\n📋 Competições a processar: {[c['nome'] for c in competicoes]}")

        for idx, competicao in enumerate(competicoes, 1):
            nome_comp = competicao['nome']
            url_comp = competicao['url']
            slug = sanitizar_nome(nome_comp)
            nome_arquivo = f"{slug}_{timestamp}.xlsx"

            print("\n" + "#" * 60)
            print(f" [{idx}/{len(competicoes)}] {nome_comp}")
            print(f" Arquivo: {nome_arquivo}")
            print("#" * 60)

            if not navegar_ate_odds(driver, url=url_comp):
                print(f"⚠️  Pulando competição: {nome_comp}")
                continue

            workbook = Workbook()
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

            extrair_competicao(driver, workbook, url_comp)

            if workbook.sheetnames:
                try:
                    workbook.save(nome_arquivo)
                    arquivos_salvos.append(nome_arquivo)
                    print(f"✅ Arquivo salvo: {nome_arquivo}  (abas: {', '.join(workbook.sheetnames)})")
                except Exception as e:
                    print(f"❌ Erro ao salvar {nome_arquivo}: {e}")
            else:
                print(f"⚠️  Nenhuma aba gerada para '{nome_comp}', arquivo não salvo.")

    except Exception as e:
        print(f"❌ Erro inesperado: {e}")
    finally:
        print("\n🔒 Fechando navegador...")
        driver.quit()

    print("\n" + "=" * 60)
    print(" RESUMO FINAL")
    print("=" * 60)
    if arquivos_salvos:
        print(f"✨ {len(arquivos_salvos)} arquivo(s) gerado(s):")
        for arq in arquivos_salvos:
            print(f"   • {arq}")
    else:
        print("⚠️  Nenhum arquivo foi gerado.")
    print("=" * 60)






if __name__ == "__main__":
    main()
